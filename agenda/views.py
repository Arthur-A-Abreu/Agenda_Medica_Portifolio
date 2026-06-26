from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from medicos.models import Medico
from plantoes.models import Plantao
from django.db.models import Count, Q
from django.http import HttpResponse
from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dateutil.relativedelta import relativedelta

MONTHS_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

@login_required
def agenda_medicos(request):
    """Lista todos os médicos para seleção da agenda com contagem do mês selecionado."""
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            year, month = today.year, today.month
    except (ValueError, TypeError):
        year, month = today.year, today.month

    medicos = Medico.objects.annotate(
        total_plantoes=Count('plantoes', filter=Q(plantoes__data__year=year, plantoes__data__month=month)),
        total_dia=Count('plantoes', filter=Q(plantoes__periodo='D', plantoes__data__year=year, plantoes__data__month=month)),
        total_manha=Count('plantoes', filter=Q(plantoes__periodo='M', plantoes__data__year=year, plantoes__data__month=month)),
        total_tarde=Count('plantoes', filter=Q(plantoes__periodo='T', plantoes__data__year=year, plantoes__data__month=month)),
        total_noite=Count('plantoes', filter=Q(plantoes__periodo='N', plantoes__data__year=year, plantoes__data__month=month))
    ).order_by('nome')

    # Navegação de meses
    months_nav = []
    start_date = date(year, month, 1)
    for i in range(-4, 5):
        m_dt = start_date + relativedelta(months=i)
        months_nav.append({
            'year': m_dt.year,
            'month': m_dt.month,
            'name': f"{MONTHS_PT[m_dt.month][:3].upper()} {str(m_dt.year)[2:]}",
            'full_name': f"{MONTHS_PT[m_dt.month]} {m_dt.year}",
            'active': m_dt.year == year and m_dt.month == month
        })

    current_month_name = f"{MONTHS_PT[month]} {year}"

    return render(request, 'agenda_medicos.html', {
        'medicos': medicos, 
        'mes_atual': month, 
        'ano_atual': year,
        'current_month_name': current_month_name,
        'months_nav': months_nav
    })

@login_required
def agenda_detalhes(request, pk):
    """Exibe a agenda detalhada de um médico específico."""
    medico = get_object_or_404(Medico, pk=pk)
    
    # Validação de permissão: Apenas Admin pode ver a agenda de outros
    if not request.user.is_superuser:
        if not hasattr(request.user, 'perfil_medico') or request.user.perfil_medico.pk != medico.pk:
            return redirect('agenda_medicos')
    
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            year, month = today.year, today.month
    except (ValueError, TypeError):
        year, month = today.year, today.month

    # Pegamos todos os plantões do médico
    plantoes_qs = medico.plantoes.all().order_by('data', 'periodo')
    
    # Agrupamos os plantões por data
    agenda_dict = defaultdict(list)
    for p in plantoes_qs:
        agenda_dict[p.data].append(p)
    
    # Ordenamos o dicionário por data
    agenda_ordenada = sorted(agenda_dict.items())
    
    agenda_passada = []
    agenda_futura = []
    
    for d, ps in agenda_ordenada:
        if d < today:
            agenda_passada.append((d, ps))
        else:
            agenda_futura.append((d, ps))
            
    # Navegação de meses
    months_nav = []
    start_date = date(year, month, 1)
    for i in range(-4, 5):
        m_dt = start_date + relativedelta(months=i)
        months_nav.append({
            'year': m_dt.year,
            'month': m_dt.month,
            'name': f"{MONTHS_PT[m_dt.month][:3].upper()} {str(m_dt.year)[2:]}",
            'full_name': f"{MONTHS_PT[m_dt.month]} {m_dt.year}",
            'active': m_dt.year == year and m_dt.month == month
        })

    current_month_name = f"{MONTHS_PT[month]} {year}"

    context = {
        'medico': medico,
        'agenda_futura': agenda_futura,
        'agenda_passada': agenda_passada[::-1], # Histórico (mais recente primeiro)
        'total_plantoes': plantoes_qs.filter(data__year=year, data__month=month).count(),
        'total_dia': plantoes_qs.filter(periodo='D', data__year=year, data__month=month).count(),
        'total_manha': plantoes_qs.filter(periodo='M', data__year=year, data__month=month).count(),
        'total_tarde': plantoes_qs.filter(periodo='T', data__year=year, data__month=month).count(),
        'total_noite': plantoes_qs.filter(periodo='N', data__year=year, data__month=month).count(),
        'today': today,
        'mes_atual': month,
        'ano_atual': year,
        'current_month_name': current_month_name,
        'months_nav': months_nav,
    }
    return render(request, 'agenda_detalhes.html', context)

@login_required
def exportar_agenda_excel(request, pk):
    """Exporta a agenda de plantões de um médico para Excel."""
    medico = get_object_or_404(Medico, pk=pk)
    
    if not request.user.is_superuser:
        return redirect('agenda_medicos')

    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            year, month = today.year, today.month
    except (ValueError, TypeError):
        year, month = today.year, today.month

    plantoes = medico.plantoes.filter(data__year=year, data__month=month).order_by('data', 'periodo')
    
    # Criar um workbook do Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Agenda {medico.nome[:15]}" # O título da planilha tem limite de tamanho
    
    # Cabeçalhos
    ws.append(['Data', 'Período', 'Status', 'Médico'])
    
    total_dia = 0
    total_manha = 0
    total_tarde = 0
    total_noite = 0
    total_concluidos = 0
    total_faltas = 0
    total_pendentes = 0
    total_concluidos_dia = 0
    total_concluidos_manha = 0
    total_concluidos_tarde = 0
    total_concluidos_noite = 0
    
    for p in plantoes:
        if p.periodo == 'D':
            total_dia += 1
        elif p.periodo == 'M':
            total_manha += 1
        elif p.periodo == 'T':
            total_tarde += 1
        elif p.periodo == 'N':
            total_noite += 1
            
        if p.status == 'C':
            total_concluidos += 1
            if p.periodo == 'D':
                total_concluidos_dia += 1
            elif p.periodo == 'M':
                total_concluidos_manha += 1
            elif p.periodo == 'T':
                total_concluidos_tarde += 1
            elif p.periodo == 'N':
                total_concluidos_noite += 1
        elif p.status == 'F':
            total_faltas += 1
        else:
            total_pendentes += 1
            
        ws.append([
            p.data.strftime('%d/%m/%Y'),
            p.get_periodo_display(),
            p.get_status_display(),
            medico.nome
        ])
    
    # Estilos Base
    header_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin', color='D1D5DB'), 
                         right=Side(style='thin', color='D1D5DB'), 
                         top=Side(style='thin', color='D1D5DB'), 
                         bottom=Side(style='thin', color='D1D5DB'))

    # Formatar Cabeçalho da Lista
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Formatar Dados da Lista
    for row in ws.iter_rows(min_row=2, max_row=len(plantoes)+1, max_col=4):
        for cell in row:
            cell.alignment = align_center
            cell.border = thin_border
    
    # Adicionar linha em branco e os totais
    ws.append([])
    
    resumo_titulo_row = len(plantoes) + 3
    ws.append(['Resumo de Plantões:'])
    ws.cell(row=resumo_titulo_row, column=1).font = Font(bold=True, size=12)
    
    resumo_dados = [
        ['Total', plantoes.count()],
        ['Dia', total_dia],
        ['Manhã', total_manha],
        ['Tarde', total_tarde],
        ['Noite', total_noite],
        ['Concluídos', total_concluidos],
        ['Faltas', total_faltas],
        ['Pendentes', total_pendentes],
        ['Concluídos (Dia)', total_concluidos_dia],
        ['Concluídos (Manhã)', total_concluidos_manha],
        ['Concluídos (Tarde)', total_concluidos_tarde],
        ['Concluídos (Noite)', total_concluidos_noite]
    ]
    
    for r in resumo_dados:
        ws.append(r)
        
    # Formatar Resumo
    for row in ws.iter_rows(min_row=resumo_titulo_row+1, max_row=resumo_titulo_row+len(resumo_dados), max_col=2):
        row[0].font = Font(bold=True)
        row[0].alignment = align_left
        row[1].alignment = align_center
    
    # Ajustar a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Pega a letra da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Preparar a resposta HTTP com o Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="agenda_{medico.nome.replace(" ", "_").lower()}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def exportar_agenda_geral_excel(request):
    """Exporta a agenda de plantões de todos os médicos para Excel."""
    
    if not request.user.is_superuser:
        return redirect('agenda_medicos')

    # Criar um workbook do Excel
    wb = openpyxl.Workbook()
    
    # Estilos Base
    header_font = Font(color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin', color='D1D5DB'), 
                         right=Side(style='thin', color='D1D5DB'), 
                         top=Side(style='thin', color='D1D5DB'), 
                         bottom=Side(style='thin', color='D1D5DB'))

    # Planilha 1: Resumo por Médico
    ws_resumo = wb.active
    ws_resumo.title = "Resumo por Médico"
    ws_resumo.append(['Médico', 'CRM', 'Total', 'Dia', 'Manhã', 'Tarde', 'Noite', 'Concluídos', 'Faltas', 'Pendentes', 'Concluídos (D)', 'Concluídos (M)', 'Concluídos (T)', 'Concluídos (N)'])
    
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            year, month = today.year, today.month
    except (ValueError, TypeError):
        year, month = today.year, today.month

    medicos = Medico.objects.annotate(
        total_plantoes=Count('plantoes', filter=Q(plantoes__data__year=year, plantoes__data__month=month)),
        total_dia=Count('plantoes', filter=Q(plantoes__periodo='D', plantoes__data__year=year, plantoes__data__month=month)),
        total_manha=Count('plantoes', filter=Q(plantoes__periodo='M', plantoes__data__year=year, plantoes__data__month=month)),
        total_tarde=Count('plantoes', filter=Q(plantoes__periodo='T', plantoes__data__year=year, plantoes__data__month=month)),
        total_noite=Count('plantoes', filter=Q(plantoes__periodo='N', plantoes__data__year=year, plantoes__data__month=month)),
        total_concluidos=Count('plantoes', filter=Q(plantoes__status='C', plantoes__data__year=year, plantoes__data__month=month)),
        total_faltas=Count('plantoes', filter=Q(plantoes__status='F', plantoes__data__year=year, plantoes__data__month=month)),
        total_pendentes=Count('plantoes', filter=Q(plantoes__status='P', plantoes__data__year=year, plantoes__data__month=month)),
        total_concluidos_dia=Count('plantoes', filter=Q(plantoes__status='C', plantoes__periodo='D', plantoes__data__year=year, plantoes__data__month=month)),
        total_concluidos_manha=Count('plantoes', filter=Q(plantoes__status='C', plantoes__periodo='M', plantoes__data__year=year, plantoes__data__month=month)),
        total_concluidos_tarde=Count('plantoes', filter=Q(plantoes__status='C', plantoes__periodo='T', plantoes__data__year=year, plantoes__data__month=month)),
        total_concluidos_noite=Count('plantoes', filter=Q(plantoes__status='C', plantoes__periodo='N', plantoes__data__year=year, plantoes__data__month=month))
    ).order_by('nome')
    
    for m in medicos:
        ws_resumo.append([
            m.nome,
            m.crm,
            m.total_plantoes,
            m.total_dia,
            m.total_manha,
            m.total_tarde,
            m.total_noite,
            m.total_concluidos,
            m.total_faltas,
            m.total_pendentes,
            m.total_concluidos_dia,
            m.total_concluidos_manha,
            m.total_concluidos_tarde,
            m.total_concluidos_noite
        ])
        
    # Formatar Cabeçalho e Dados do Resumo
    for row in ws_resumo.iter_rows(min_row=1, max_row=len(medicos)+1, max_col=14):
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
    # Ajustar colunas do resumo
    for col in ws_resumo.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_resumo.column_dimensions[column].width = (max_length + 2)

    # Planilha 2: Todos os Plantões (Lista detalhada)
    ws_detalhes = wb.create_sheet(title="Todos os Plantões")
    ws_detalhes.append(['Data', 'Período', 'Status', 'Médico', 'CRM'])
    
    plantoes = Plantao.objects.filter(data__year=year, data__month=month).select_related('medico').order_by('data', 'periodo')
    
    for p in plantoes:
        ws_detalhes.append([
            p.data.strftime('%d/%m/%Y'),
            p.get_periodo_display(),
            p.get_status_display(),
            p.medico.nome,
            p.medico.crm
        ])
        
    # Formatar Cabeçalho e Dados dos Detalhes
    for row in ws_detalhes.iter_rows(min_row=1, max_row=len(plantoes)+1, max_col=5):
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
    # Ajustar colunas dos detalhes
    for col in ws_detalhes.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_detalhes.column_dimensions[column].width = (max_length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="agenda_geral_todos_medicos.xlsx"'
    
    wb.save(response)
    return response
